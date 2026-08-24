"""
Gera a planilha final (.xlsx) a partir dos resultados do motor de análise.

Estrutura em abas separadas:
    - "Resumo"      — visão geral de todos os produtos, uma linha por item
    - "ICMS"        — detalhe de benefício e substituição tributária de ICMS
    - "PIS-COFINS"  — detalhe de benefício e regime monofásico de PIS/COFINS
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


FONTE_CABECALHO = Font(bold=True, color="FFFFFF")
FUNDO_CABECALHO = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
FUNDO_SIM = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FUNDO_NAO = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
FUNDO_ALERTA = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FUNDO_INCOERENTE = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
FUNDO_REGIME_ESPECIAL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")


def _estilizar_cabecalho(ws, cabecalho):
    ws.append(cabecalho)
    for col_idx in range(1, len(cabecalho) + 1):
        celula = ws.cell(row=1, column=col_idx)
        celula.font = FONTE_CABECALHO
        celula.fill = FUNDO_CABECALHO
        celula.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"


def _ajustar_larguras(ws, larguras):
    for idx, largura in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = largura


def _detalhe_beneficios(lista_beneficios):
    if not lista_beneficios:
        return "-"
    return " | ".join(
        f"{b.tipo_beneficio.upper()} - {b.norma_titulo}" +
        (f" [{b.condicoes}]" if b.condicoes else "")
        for b in lista_beneficios
    )


def _detalhe_st(detalhe_st):
    if not detalhe_st:
        return "-"
    return " | ".join(
        f"CEST {s.cest or '-'} / UF {s.uf or '-'}" +
        (f" / MVA {s.mva_percentual}%" if s.mva_percentual is not None else "") +
        (f" - {s.norma_titulo}" if s.norma_titulo else "")
        for s in detalhe_st
    )


def _detalhe_monofasico(regime_monofasico):
    if not regime_monofasico:
        return "-"
    rm = regime_monofasico
    if rm.aliquota_zero:
        return f"Alíquota ZERO (revenda) - código {rm.codigo_sped} - {rm.origem_tabela}"
    return (
        f"PIS {rm.aliquota_pis}% / COFINS {rm.aliquota_cofins}% "
        f"- código {rm.codigo_sped} - {rm.origem_tabela}"
    )


def _preencher_aba_resumo(ws, resultados, pareceres_ia):
    cabecalho = [
        "Descrição do Produto", "NCM", "NCM Válido (TIPI)",
        "Tem Benefício ICMS", "Sujeito a ST-ICMS",
        "Tem Benefício PIS/COFINS", "Regime Monofásico PIS/COFINS",
        "Descrição TIPI", "Alíquota IPI",
        "Classificação Coerente (IA)", "Alerta",
    ]
    _estilizar_cabecalho(ws, cabecalho)

    for r in resultados:
        chave_ia = f"{r.ncm}|{r.descricao_produto}"
        parecer = pareceres_ia.get(chave_ia)
        coerente_txt = ("Sim" if parecer.classificacao_coerente else "NÃO") if parecer else "-"

        linha = [
            r.descricao_produto, r.ncm, "Sim" if r.ncm_valido else "Não",
            "Sim" if r.tem_beneficio_icms else "Não",
            "Sim" if r.sujeito_st_icms else "Não",
            "Sim" if r.tem_beneficio_piscofins else "Não",
            "Sim" if r.regime_monofasico else "Não",
            r.tipi_descricao or "-", r.tipi_aliquota_ipi or "-",
            coerente_txt, r.alerta or "-",
        ]
        ws.append(linha)
        linha_atual = ws.max_row
        ws.cell(row=linha_atual, column=4).fill = FUNDO_SIM if r.tem_beneficio_icms else FUNDO_NAO
        ws.cell(row=linha_atual, column=5).fill = FUNDO_REGIME_ESPECIAL if r.sujeito_st_icms else FUNDO_NAO
        ws.cell(row=linha_atual, column=6).fill = FUNDO_SIM if r.tem_beneficio_piscofins else FUNDO_NAO
        ws.cell(row=linha_atual, column=7).fill = FUNDO_REGIME_ESPECIAL if r.regime_monofasico else FUNDO_NAO

        if r.alerta:
            for col_idx in range(1, len(cabecalho) + 1):
                ws.cell(row=linha_atual, column=col_idx).fill = FUNDO_ALERTA
        if parecer and not parecer.classificacao_coerente:
            for col_idx in range(1, len(cabecalho) + 1):
                ws.cell(row=linha_atual, column=col_idx).fill = FUNDO_INCOERENTE

    _ajustar_larguras(ws, [35, 12, 14, 14, 14, 16, 20, 35, 12, 16, 35])


def _preencher_aba_icms(ws, resultados):
    cabecalho = [
        "Descrição do Produto", "NCM", "Tem Benefício ICMS",
        "Detalhe do Benefício ICMS", "Sujeito a ST-ICMS", "Detalhe da ST-ICMS",
    ]
    _estilizar_cabecalho(ws, cabecalho)

    for r in resultados:
        linha = [
            r.descricao_produto, r.ncm,
            "Sim" if r.tem_beneficio_icms else "Não",
            _detalhe_beneficios(r.beneficios_icms),
            "Sim" if r.sujeito_st_icms else "Não",
            _detalhe_st(r.detalhe_st),
        ]
        ws.append(linha)
        linha_atual = ws.max_row
        ws.cell(row=linha_atual, column=3).fill = FUNDO_SIM if r.tem_beneficio_icms else FUNDO_NAO
        ws.cell(row=linha_atual, column=5).fill = FUNDO_REGIME_ESPECIAL if r.sujeito_st_icms else FUNDO_NAO

    _ajustar_larguras(ws, [35, 12, 16, 50, 14, 50])


def _preencher_aba_piscofins(ws, resultados):
    cabecalho = [
        "Descrição do Produto", "NCM", "Tem Benefício PIS/COFINS",
        "Detalhe do Benefício PIS/COFINS", "Regime Monofásico", "Detalhe do Regime Monofásico",
    ]
    _estilizar_cabecalho(ws, cabecalho)

    for r in resultados:
        linha = [
            r.descricao_produto, r.ncm,
            "Sim" if r.tem_beneficio_piscofins else "Não",
            _detalhe_beneficios(r.beneficios_piscofins),
            "Sim" if r.regime_monofasico else "Não",
            _detalhe_monofasico(r.regime_monofasico),
        ]
        ws.append(linha)
        linha_atual = ws.max_row
        ws.cell(row=linha_atual, column=3).fill = FUNDO_SIM if r.tem_beneficio_piscofins else FUNDO_NAO
        ws.cell(row=linha_atual, column=5).fill = FUNDO_REGIME_ESPECIAL if r.regime_monofasico else FUNDO_NAO

    _ajustar_larguras(ws, [35, 12, 18, 50, 16, 50])


def _preencher_aba_ia(ws, resultados, pareceres_ia):
    cabecalho = [
        "Descrição do Produto", "NCM", "Descrição TIPI",
        "Classificação Coerente", "Confiança", "Parecer da IA", "NCM Sugerido",
    ]
    _estilizar_cabecalho(ws, cabecalho)

    for r in resultados:
        chave_ia = f"{r.ncm}|{r.descricao_produto}"
        parecer = pareceres_ia.get(chave_ia)
        if not parecer:
            continue
        linha = [
            r.descricao_produto, r.ncm, r.tipi_descricao or "-",
            "Sim" if parecer.classificacao_coerente else "NÃO",
            parecer.nivel_confianca, parecer.justificativa, parecer.ncm_sugerido or "-",
        ]
        ws.append(linha)
        if not parecer.classificacao_coerente:
            for col_idx in range(1, len(cabecalho) + 1):
                ws.cell(row=ws.max_row, column=col_idx).fill = FUNDO_INCOERENTE

    _ajustar_larguras(ws, [35, 12, 35, 18, 12, 50, 16])


def gerar_planilha_resultado(resultados: list, caminho_saida: str, pareceres_ia: dict = None):
    """
    resultados: lista de ResultadoItem (do motor_analise.py)
    caminho_saida: caminho do .xlsx a ser gerado
    pareceres_ia: dict opcional {ncm+descricao: ParecerIA}, do classificador_ia.py.

    Gera a planilha em abas separadas: Resumo, ICMS, PIS-COFINS e (se houver
    pareceres de IA) uma aba extra "Classificação IA".
    """
    pareceres_ia = pareceres_ia or {}

    wb = openpyxl.Workbook()
    ws_resumo = wb.active
    ws_resumo.title = "Resumo"
    _preencher_aba_resumo(ws_resumo, resultados, pareceres_ia)

    ws_icms = wb.create_sheet("ICMS")
    _preencher_aba_icms(ws_icms, resultados)

    ws_piscofins = wb.create_sheet("PIS-COFINS")
    _preencher_aba_piscofins(ws_piscofins, resultados)

    if pareceres_ia:
        ws_ia = wb.create_sheet("Classificação IA")
        _preencher_aba_ia(ws_ia, resultados, pareceres_ia)

    wb.save(caminho_saida)
    return caminho_saida


if __name__ == "__main__":
    from motor_analise import analisar_lote

    tipi_exemplo = {
        "73181500": {"descricao": "Parafusos de ferro fundido, ferro ou aço", "aliquota_ipi": "5"},
        "30049099": {"descricao": "Outros medicamentos", "aliquota_ipi": "NT"},
        "22011000": {"descricao": "Águas minerais e águas gaseificadas", "aliquota_ipi": "0"},
    }
    beneficios_exemplo = [
        {"ncm": "3004", "ncm_prefixo": True, "norma_titulo": "Convênio ICMS 87/2002", "tributo": "ICMS",
         "tipo_beneficio": "isencao", "condicoes": "medicamentos de uso humano", "vigencia_fim": None},
        {"ncm": "3004", "ncm_prefixo": True, "norma_titulo": "Lei 10.147/2000", "tributo": "PIS/COFINS",
         "tipo_beneficio": "aliquota zero", "condicoes": "medicamentos relacionados em ato do Executivo", "vigencia_fim": None},
    ]
    tabela_monofasico = [
        {"ncm": "22011000", "codigo": "822", "descricao_produto": "Águas Minerais >= 10L",
         "aliquota_pis": "0,00", "aliquota_cofins": "0,00", "aliquota_zero": True, "origem_tabela": "Tabela 4.3.11"},
    ]
    tabela_st = [
        {"ncm": "73181500", "ncm_prefixo": False, "cest": "10.123.00", "uf": "GO",
         "mva_percentual": 40.0, "norma_titulo": "Protocolo ICMS 41/2008", "observacoes": None},
    ]
    itens_exemplo = [
        {"descricao_produto": "PARAFUSO DE ACO INOX M6", "ncm": "73181500"},
        {"descricao_produto": "MEDICAMENTO GENERICO XYZ 500MG", "ncm": "30049099"},
        {"descricao_produto": "AGUA MINERAL 10L GALAO", "ncm": "22011000"},
        {"descricao_produto": "NOTEBOOK DELL I7 16GB", "ncm": "73181500"},
    ]

    from classificador_ia import ParecerIA
    resultados = analisar_lote(itens_exemplo, tipi_exemplo, beneficios_exemplo, tabela_monofasico, tabela_st)
    pareceres_simulados = {
        "73181500|NOTEBOOK DELL I7 16GB": ParecerIA(
            ncm="73181500", descricao_produto="NOTEBOOK DELL I7 16GB",
            classificacao_coerente=False, nivel_confianca="alta",
            justificativa="Notebook é equipamento de informática, incompatível com parafusos.",
            ncm_sugerido="84713012"),
    }
    caminho = gerar_planilha_resultado(resultados, "/home/claude/sistema-fiscal/data/teste_abas.xlsx", pareceres_simulados)
    print(f"Planilha gerada em: {caminho}")

    import openpyxl as ox
    wb_check = ox.load_workbook(caminho)
    print("Abas geradas:", wb_check.sheetnames)
