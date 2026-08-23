"""
Parser da Tabela TIPI (Receita Federal).

A Receita disponibiliza a TIPI em .xlsx e .pdf. O layout do xlsx oficial
costuma vir com colunas tipo: NCM | DESCRIÇÃO | ALÍQUOTA (%).
Como o layout pode variar levemente entre versões, este parser tenta
detectar automaticamente as colunas pelo cabeçalho.

Uso:
    from parser_tipi import processar_tipi_xlsx

    linhas = processar_tipi_xlsx("/caminho/tipi.xlsx", versao_fonte="TIPI 2022 - ADE 001-2026")
"""

import re
import openpyxl
from dataclasses import dataclass, asdict
from typing import Optional


@dataclass
class RegistroTipi:
    ncm: str
    descricao: str
    aliquota_ipi: Optional[str]
    capitulo: Optional[str]
    ex_tarifario: Optional[str]
    versao_fonte: str


# Regex pra validar formato de NCM: 8 dígitos, podendo vir com pontos (0000.00.00)
NCM_REGEX = re.compile(r"^\d{4}\.?\d{2}\.?\d{2}$")


def _limpar_ncm(valor) -> Optional[str]:
    """Remove pontuação do NCM e valida se tem 8 dígitos."""
    if valor is None:
        return None
    texto = str(valor).strip().replace(".", "").replace(" ", "")
    if texto.isdigit() and len(texto) == 8:
        return texto
    return None


def _detectar_colunas(planilha) -> dict:
    """
    Varre as primeiras linhas procurando o cabeçalho e mapeia os índices
    das colunas de interesse (NCM, descrição, alíquota, ex-tarifário).
    """
    mapa_colunas = {}
    for num_linha, linha in enumerate(planilha.iter_rows(min_row=1, max_row=15), start=1):
        valores = [str(c.value).strip().upper() if c.value else "" for c in linha]
        for idx, valor in enumerate(valores):
            if "NCM" in valor and "ncm" not in mapa_colunas:
                mapa_colunas["ncm"] = idx
            elif ("DESCRI" in valor) and "descricao" not in mapa_colunas:
                mapa_colunas["descricao"] = idx
            elif ("ALÍQUOTA" in valor or "ALIQUOTA" in valor or valor == "IPI") and "aliquota" not in mapa_colunas:
                mapa_colunas["aliquota"] = idx
            elif "EX" == valor and "ex" not in mapa_colunas:
                mapa_colunas["ex"] = idx

        if "ncm" in mapa_colunas and "descricao" in mapa_colunas:
            mapa_colunas["linha_cabecalho"] = num_linha
            break

    return mapa_colunas


def processar_tipi_xlsx(caminho_xlsx: str, versao_fonte: str = "TIPI") -> list:
    """
    Lê o xlsx oficial da TIPI e retorna uma lista de RegistroTipi.

    Trata o padrão da TIPI onde a descrição de um capítulo/posição
    aparece numa linha e o NCM completo (8 dígitos) só aparece nas
    linhas de subitem — linhas sem NCM válido de 8 dígitos são puladas
    (são apenas texto de agrupamento/nota).
    """
    wb = openpyxl.load_workbook(caminho_xlsx, data_only=True)
    planilha = wb.active

    colunas = _detectar_colunas(planilha)
    if "ncm" not in colunas or "descricao" not in colunas:
        raise ValueError(
            "Não consegui identificar as colunas de NCM/Descrição automaticamente. "
            "Verifique o layout do arquivo — pode ter mudado em relação ao esperado."
        )

    linha_inicio = colunas.get("linha_cabecalho", 1) + 1
    registros = []

    for linha in planilha.iter_rows(min_row=linha_inicio, values_only=False):
        valores = [c.value for c in linha]

        ncm_bruto = valores[colunas["ncm"]] if colunas["ncm"] < len(valores) else None
        descricao_bruta = valores[colunas["descricao"]] if colunas["descricao"] < len(valores) else None
        aliquota_bruta = valores[colunas.get("aliquota", -1)] if colunas.get("aliquota", -1) >= 0 and colunas.get("aliquota", -1) < len(valores) else None
        ex_bruto = valores[colunas.get("ex", -1)] if colunas.get("ex", -1) >= 0 and colunas.get("ex", -1) < len(valores) else None

        ncm = _limpar_ncm(ncm_bruto)

        if not ncm or not descricao_bruta:
            continue  # linha de agrupamento/nota (capítulo/posição), sem NCM completo — pula

        # O capítulo é sempre determinístico a partir dos 2 primeiros dígitos
        # do próprio NCM de 8 dígitos — mais confiável do que tentar detectar
        # o texto "Capítulo N" no arquivo (que pode vir formatado de forma
        # inconsistente entre versões da planilha).
        registros.append(RegistroTipi(
            ncm=ncm,
            descricao=str(descricao_bruta).strip(),
            aliquota_ipi=str(aliquota_bruta).strip() if aliquota_bruta is not None else None,
            capitulo=ncm[:2],
            ex_tarifario=str(ex_bruto).strip() if ex_bruto else None,
            versao_fonte=versao_fonte,
        ))

    return registros


def gerar_relatorio_carga(registros: list) -> dict:
    """Estatísticas rápidas pra conferência após a carga."""
    return {
        "total_ncms": len(registros),
        "ncms_nao_tributados": sum(1 for r in registros if r.aliquota_ipi and r.aliquota_ipi.upper() == "NT"),
        "ncms_com_ex_tarifario": sum(1 for r in registros if r.ex_tarifario),
        "capitulos_distintos": len(set(r.capitulo for r in registros if r.capitulo)),
    }


if __name__ == "__main__":
    import sys
    import json

    if len(sys.argv) < 2:
        print("Uso: python parser_tipi.py caminho/para/tipi.xlsx")
        sys.exit(1)

    registros = processar_tipi_xlsx(sys.argv[1], versao_fonte="TIPI 2022 - Atualizada ADE 001-2026")
    print(json.dumps(gerar_relatorio_carga(registros), indent=2, ensure_ascii=False))
    print("\nExemplos:")
    for r in registros[:5]:
        print(asdict(r))
